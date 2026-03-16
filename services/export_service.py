from __future__ import annotations

import csv
import io
from datetime import datetime, timezone
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _build_rows(ctx: dict) -> tuple[list[str], list[list[Any]]]:
    categories = ctx["categories"]
    students = ctx["students"]

    headers = [
        "Rank", "Full Name", "Student Code", "Gender", "Badge",
        "Avg Grade", "Letter Grade", "Attendance Rate (%)",
        "Present", "Late", "Absent", "Excused",
    ]
    for cat in categories:
        headers.append(cat["name"])

    rows: list[list[Any]] = []
    for s in students:
        att = s["attendance"]
        row: list[Any] = [
            s["rank"],
            s["full_name"],
            s["student_code"],
            s["gender"],
            s["badge"],
            s["avg_grade"] if s["avg_grade"] is not None else "",
            s["letter_grade"],
            f'{att["rate"] * 100:.1f}',
            att["present"],
            att["late"],
            att["absent"],
            att["excused"],
        ]
        for cat in categories:
            cid = str(cat["id"])
            cat_data = s["grades_by_category"].get(cid, {})
            avg = cat_data.get("avg")
            row.append(avg if avg is not None else "")
        rows.append(row)

    return headers, rows


def export_csv(ctx: dict) -> tuple[str, str]:
    headers, rows = _build_rows(ctx)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(rows)

    classroom_name = ctx["classroom"].get("name", "report").replace(" ", "_")
    date_str = datetime.now(timezone.utc).strftime("%Y%m%d")
    filename = f"{classroom_name}_report_{date_str}.csv"
    return output.getvalue(), filename


def export_excel(ctx: dict) -> tuple[bytes, str]:
    headers, rows = _build_rows(ctx)

    wb = openpyxl.Workbook()
    ws = wb.active
    classroom_name = ctx["classroom"].get("name", "Report")
    ws.title = "Class Report"

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f'{classroom_name} — Class Report'
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    meta_cell = ws.cell(row=2, column=1)
    meta_cell.value = (
        f'Subject: {ctx["classroom"].get("subject", "")}  |  '
        f'Room: {ctx["classroom"].get("room", "")}  |  '
        f'Sessions: {ctx["total_sessions"]}  |  '
        f'Generated: {datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")}'
    )
    meta_cell.font = Font(italic=True, size=10)
    meta_cell.alignment = Alignment(horizontal="center")

    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row_idx, row_data in enumerate(rows, start=4):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F5F8FF", end_color="F5F8FF", fill_type="solid")

    grade_col = 7
    grade_colors = {"A": "C6EFCE", "B": "DDEBF7", "C": "FFEB9C", "D": "FCE4D6", "F": "FFC7CE"}
    for row_idx in range(4, 4 + len(rows)):
        cell = ws.cell(row=row_idx, column=grade_col)
        color = grade_colors.get(str(cell.value))
        if color:
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.font = Font(bold=True)

    summary_row = 4 + len(rows)
    ws.cell(row=summary_row, column=1, value="CLASS AVG").font = Font(bold=True)
    ws.cell(row=summary_row, column=6, value=ctx.get("class_avg") or "").font = Font(bold=True)
    ws.cell(row=summary_row, column=8, value=f'{ctx["class_attendance_rate"] * 100:.1f}').font = Font(bold=True)

    col_widths = [6, 24, 14, 10, 12, 10, 12, 16, 8, 8, 8, 8]
    for cat in ctx["categories"]:
        col_widths.append(max(12, len(cat["name"]) + 2))
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A4"
    ws.row_dimensions[3].height = 36

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    date_str = datetime.now(timezone.utc).strftime("%Y%m%d")
    safe_name = classroom_name.replace(" ", "_")
    filename = f"{safe_name}_report_{date_str}.xlsx"
    return output.getvalue(), filename
